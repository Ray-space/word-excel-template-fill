import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from parse_exam_questions import (
    Exporter,
    Parser,
    build_template_layout,
    normalize_answer,
    validate_template_layout,
)


class ParseExamQuestionsTests(unittest.TestCase):
    def setUp(self) -> None:
        self.parser = Parser()
        self.exporter = Exporter(answer_separator=",")
        # 与导题模板列位一致（至少 17 列），便于按模板表头映射
        self.columns = [
            "选项（1单选 2多选 3判断）",
            "题干",
            "选项A",
            "选项B",
            "选项C",
            "选项D",
            "选项E",
            "选项F",
            "选项G",
            "选项H",
            "答案",
            "题目解析",
            "试题难度",
            "知识点ID",
            "知识点名称",
            "标签",
            "指标",
        ]
        self.layout = build_template_layout(self.columns)

    def _row_dict(self, question) -> dict:
        lst = self.exporter.build_row(question, "企业所得税", self.layout)
        return self.exporter.row_as_dict(self.layout, lst)

    def test_normal_question(self) -> None:
        lines = [
            "1、根据规定，下列说法正确的是（ ）",
            "A. 选项一",
            "B. 选项二",
            "C. 选项三",
            "D. 选项四",
            "答案：D",
            "题目解析：这是解析内容，长度足够。难度：一般",
            "知识点：24BWQ040186减免所得税",
        ]
        questions, errors = self.parser.parse_questions(lines)
        self.assertEqual(1, len(questions))
        self.assertEqual(0, len(errors))
        row = self._row_dict(questions[0])
        self.assertEqual("1", row["选项（1单选 2多选 3判断）"])
        self.assertEqual("D", row["答案"])
        self.assertEqual("一般", row["试题难度"])
        self.assertEqual("24BWQ040186减免所得税", row["知识点名称"])
        self.assertEqual("企业所得税", row["标签"])

    def test_missing_option_question(self) -> None:
        lines = [
            "2、这是一道缺少选项的题（ ）",
            "A. 正确",
            "答案：A",
            "题目解析：解析内容仍然存在。难度：简单",
            "知识点：税收基础",
        ]
        questions, errors = self.parser.parse_questions(lines)
        self.assertEqual(1, len(questions))
        self.assertEqual(0, len(errors))
        row = self._row_dict(questions[0])
        self.assertEqual("1", row["选项（1单选 2多选 3判断）"])
        self.assertEqual("正确", row["选项A"])
        self.assertEqual("", row["选项B"])

    def test_multiline_analysis(self) -> None:
        lines = [
            "3、这是多行解析题目（ ）",
            "A. 甲",
            "B. 乙",
            "答案：A",
            "题目解析：第一段解析。",
            "第二段解析继续说明。",
            "难度：困难",
            "知识点：企业所得税优惠",
        ]
        questions, errors = self.parser.parse_questions(lines)
        self.assertEqual(1, len(questions))
        self.assertEqual(0, len(errors))
        row = self._row_dict(questions[0])
        self.assertIn("第一段解析", row["题目解析"])
        self.assertIn("第二段解析继续说明", row["题目解析"])
        self.assertEqual("困难", row["试题难度"])

    def test_judgment_question_no_options(self) -> None:
        """判断题：无选项段落，直接答案"""
        lines = [
            "80、根据企业所得税的规定，一般企业发生的公益性捐赠支出...（ ）",
            "答案：错误",
            "题目解析：根据《企业所得税法》第九条规定...难度：一般",
        ]
        questions, errors = self.parser.parse_questions(lines)
        self.assertEqual(1, len(questions))
        self.assertEqual(0, len(errors))
        row = self._row_dict(questions[0])
        self.assertEqual("3", row["选项（1单选 2多选 3判断）"])
        self.assertEqual("", row["选项A"])
        self.assertEqual("", row["选项B"])
        self.assertEqual("B", row["答案"])

    def test_multiple_choice_six_options(self) -> None:
        """多选题：6个选项(A-F)，答案多字母带空格"""
        lines = [
            "38、根据企业所得税现行政策规定...（ ）",
            "A. 首次确认...",
            "B. 社会组织评估等级为3A以上...",
            "C. 不具有公开募捐资格...8%",
            "D. 具有公开募捐资格...10%",
            "E. 不具有公开募捐资格...12%",
            "F. F",
            "答案：A B D E",
            "题目解析：...难易程度：困难",
            "知识点：24BWQ040142成本费用扣除项目",
        ]
        questions, errors = self.parser.parse_questions(lines)
        self.assertEqual(1, len(questions))
        self.assertEqual(0, len(errors))
        row = self._row_dict(questions[0])
        self.assertEqual("2", row["选项（1单选 2多选 3判断）"])
        self.assertEqual("A,B,D,E", row["答案"])
        self.assertEqual("F", row["选项F"])

    def test_answer_format_variants(self) -> None:
        """测试多种答案格式统一"""
        test_cases = [
            ("B C", "B,C"),
            ("BC", "B,C"),
            ("B、C", "B,C"),
            ("A B D E", "A,B,D,E"),
        ]
        for raw, expected in test_cases:
            result = normalize_answer(raw)
            self.assertEqual(expected, result, f"格式'{raw}'应统一为'{expected}'")

    def test_inline_options_in_single_paragraph(self) -> None:
        lines = [
            "1、根据规定，下列说法正确的是（ ）",
            "A. 甲    B. 乙    C. 丙    D. 丁",
            "答案：B（依据：第30条）",
        ]
        questions, errors = self.parser.parse_questions(lines)
        self.assertEqual(1, len(questions))
        self.assertEqual(0, len(errors))
        row = self._row_dict(questions[0])
        self.assertEqual("甲", row["选项A"])
        self.assertEqual("乙", row["选项B"])
        self.assertEqual("丙", row["选项C"])
        self.assertEqual("丁", row["选项D"])
        self.assertEqual("B", row["答案"])
        self.assertIn("依据：第30条", row["题目解析"])

    def test_answer_line_trailing_basis_becomes_analysis(self) -> None:
        lines = [
            "1、判断题（ ）",
            "答案：错误（依据：第35条：地方一般公共预算按量入为出、收支平衡原则编制）",
        ]
        questions, errors = self.parser.parse_questions(lines)
        self.assertEqual(1, len(questions))
        self.assertEqual(0, len(errors))
        row = self._row_dict(questions[0])
        self.assertEqual("3", row["选项（1单选 2多选 3判断）"])
        self.assertEqual("B", row["答案"])
        self.assertIn("依据：第35条", row["题目解析"])

    def test_stop_parsing_after_answer_key_section(self) -> None:
        lines = [
            "1. 第一题",
            "A. 甲 B. 乙 C. 丙 D. 丁",
            "答案：A（依据：第3条）",
            "参考答案汇总",
            "1. 这是说明，不应再识别为题目",
            "2. 这也是说明",
        ]
        questions, errors = self.parser.parse_questions(lines)
        self.assertEqual(1, len(questions))
        self.assertEqual(0, len(errors))

    def test_inline_difficulty_and_knowledge_concatenated(self) -> None:
        """解析末尾「难易程度：」与「知识点：」紧连（无空格）也能正确拆分"""
        text = "解析正文。难易程度：困难知识点：24BWQ040131不征税收入（法定、特定）"
        analysis, diff, kn = self.exporter.split_analysis_meta(text)
        self.assertEqual("困难", diff)
        self.assertIn("24BWQ040131", kn)
        self.assertIn("不征税收入", kn)
        self.assertNotIn("难易程度", analysis)
        self.assertNotIn("知识点", analysis)

    def test_judgment_difficulty_and_knowledge_inline(self) -> None:
        """判断题解析后「难度：」与「知识点：」紧连在一行"""
        text = "依据规定。难度：困难知识点：24BWQ0401101房地产开发企业、建筑业"
        analysis, diff, kn = self.exporter.split_analysis_meta(text)
        self.assertEqual("困难", diff)
        self.assertIn("24BWQ0401101", kn)
        self.assertIn("房地产开发企业", kn)
        self.assertNotIn("难度", analysis)
        self.assertNotIn("知识点", analysis)

    def test_single_choice_difficulty_general_and_knowledge_inline(self) -> None:
        """单选题解析后「难度：一般」与「知识点：」紧连在一行"""
        text = "根据政策。难度：一般知识点：24BWQ040181减计收入"
        analysis, diff, kn = self.exporter.split_analysis_meta(text)
        self.assertEqual("一般", diff)
        self.assertIn("24BWQ040181", kn)
        self.assertIn("减计收入", kn)
        self.assertNotIn("难度", analysis)
        self.assertNotIn("知识点", analysis)

    def test_import_template_header_columns_from_spec(self) -> None:
        """与导题模板截图尾部表头一致（含全角括号、换行）"""
        hdrs = [""] * 10 + [
            "选项H",
            "答案（A,B 英文逗号）",
            "题目解析",
            "试题难度(简单，一般，困难)",
            "知识点ID(111,222 英文逗号)\n填写时",
            "知识点名称(知识点1,知识点2 英文逗号)",
            "标签(标签1,标签2 英文逗号)",
            "指标(指标1,指标2 英文逗号)",
        ]
        layout = build_template_layout(hdrs)
        self.assertEqual(layout.option_i.get("H"), 10)
        self.assertEqual(layout.answer_i, 11)
        self.assertEqual(layout.analysis_i, 12)
        self.assertEqual(layout.difficulty_i, 13)
        self.assertEqual(layout.kid_i, 14)
        self.assertEqual(layout.kname_i, 15)
        self.assertEqual(layout.tag_i, 16)
        self.assertEqual(layout.metric_i, 17)

    def test_validate_template_layout_ok(self) -> None:
        issues = validate_template_layout(self.layout)
        self.assertEqual([], issues)

    def test_export_row_pads_to_template_width(self) -> None:
        """行长度短于模板列数时右侧补空，避免 openpyxl 错位。"""
        exp = Exporter()
        wide = self.columns + ["", "", "备注"]
        layout = build_template_layout(wide)
        qs, _errs = self.parser.parse_questions(
            [
                "1、题干（ ）",
                "A. x",
                "B. y",
                "答案：A",
                "题目解析：说明文字足够长用于校验。难度：一般",
                "知识点：K",
            ]
        )
        row_list = exp.build_row(qs[0], "企业所得税", layout)
        self.assertEqual(len(wide), len(row_list))
        fd, out_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            exp.export_excel([row_list], wide, Path(out_path))
            wb = load_workbook(out_path, read_only=True)
            second = list(wb.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            wb.close()
            self.assertEqual(len(wide), len(second))
            self.assertIn(second[-1], (None, ""))
        finally:
            os.unlink(out_path)


if __name__ == "__main__":
    unittest.main()
