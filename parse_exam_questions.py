import argparse
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook, Workbook

OPTIONS = list("ABCDEFGH")
QUESTION_START_RE = re.compile(r"^\s*(\d+)(?:\.\s+|、\s*)(.*)$")
OPTION_RE = re.compile(r"^\s*([A-H])\s*[\.。．、)]\s*(.*)$", re.IGNORECASE)
ANSWER_RE = re.compile(r"^\s*答案\s*[:：]\s*(.+)\s*$")
ANALYSIS_START_RE = re.compile(r"^\s*(?:题目)?解析\s*[:：]\s*(.*)$")
INLINE_OPTION_RE = re.compile(r"([A-H])\s*[\.。．、)]\s*(.*?)(?=(?:\s+[A-H]\s*[\.。．、)])|$)", re.IGNORECASE)
ANSWER_HEAD_RE = re.compile(r"^\s*(正确|错误|[A-H](?:[\s,，、和及A-H]*)?)", re.IGNORECASE)
ANSWER_KEY_SECTION_RE = re.compile(r"参考答案|答案汇总|答案一览", re.IGNORECASE)
SECTION_BREAK_RE = re.compile(r"^[\W_一二三四五六七八九十多选单选判断题共每分]+$")
# 难度/难易程度：只捕获标准档位，避免「困难知识点：」连写时被整块吞掉
DIFFICULTY_LEVEL_RE = re.compile(
    r"(?:难度|难易程度)\s*[:：]?\s*(简单|一般|困难|普通|中等|较难)"
)
# 知识点：… 可能后接「难度/难易程度」或行尾；非贪婪避免误吞后续字段
KNOWLEDGE_RE = re.compile(
    r"知识点\s*[:：]\s*(.+?)(?=\s*(?:难度|难易程度)\s*[:：]|$)",
    re.DOTALL,
)
ANSWER_LETTER_RE = re.compile(r"[A-H]", re.IGNORECASE)


@dataclass
class Question:
    number: int
    raw_lines: List[str] = field(default_factory=list)
    stem_lines: List[str] = field(default_factory=list)
    options: Dict[str, str] = field(default_factory=lambda: {k: "" for k in OPTIONS})
    answer_raw: str = ""
    analysis_lines: List[str] = field(default_factory=list)


@dataclass
class ParseError:
    number: Optional[int]
    message: str
    snippet: str


@dataclass
class ValidationSummary:
    total: int
    critical: int
    warnings: int
    pass_rate: float

    def as_json(self) -> str:
        payload = {
            "total": self.total,
            "critical": self.critical,
            "warnings": self.warnings,
            "pass_rate": round(self.pass_rate, 4),
        }
        return json.dumps(payload, ensure_ascii=False)


@dataclass
class TemplateLayout:
    """与导题模板首行逐列一致，并记录各语义字段所在列下标。"""

    columns: List[str]
    type_i: Optional[int] = None
    stem_i: Optional[int] = None
    option_i: Dict[str, int] = field(default_factory=dict)
    answer_i: Optional[int] = None
    analysis_i: Optional[int] = None
    difficulty_i: Optional[int] = None
    kid_i: Optional[int] = None
    kname_i: Optional[int] = None
    tag_i: Optional[int] = None
    metric_i: Optional[int] = None


def _header_key(h: str) -> str:
    """去掉所有空白（含换行），便于与模板中带换行的表头比对。"""
    return re.sub(r"\s+", "", h)


def validate_template_layout(layout: TemplateLayout) -> List[str]:
    """导出前自检：必填语义列缺失时给出明确提示（不中断导出）。"""
    issues: List[str] = []
    if layout.type_i is None:
        issues.append("未识别「题型」列（需表头同时含单选/多选/判断等）")
    if layout.stem_i is None:
        issues.append("未识别「题干」列")
    if layout.answer_i is None:
        issues.append("未识别「答案」列")
    # 解析/难度缺失时仍可导出，但质量下降
    if layout.analysis_i is None:
        issues.append("警告: 未识别「题目解析」列，解析正文将无法写入对应列")
    if layout.difficulty_i is None:
        issues.append("警告: 未识别「试题难度/难易程度」列，难度将无法单独成列")
    if layout.tag_i is None:
        issues.append("警告: 未识别「标签」列，请检查表头是否以「标签」开头")
    # 同名/空表头：按列下标写入可规避 dict 键冲突；此处仅提示
    names = [c for c in layout.columns if str(c).strip()]
    if len(names) != len(set(names)):
        issues.append("警告: 模板首行存在重复的非空列名，后序列会覆盖前列（已用按列下标写入规避）")
    return issues


def build_template_layout(raw_columns: List[str]) -> TemplateLayout:
    # 表头字符串与模板单元格一致（不 strip），匹配规则内部再 strip / 去空白
    columns = [str(c) if c is not None else "" for c in raw_columns]
    layout = TemplateLayout(columns=columns)
    for i, raw in enumerate(columns):
        h = raw.strip()
        if not h:
            continue
        hk = _header_key(h)
        if layout.type_i is None and ("题型" in hk or ("单选" in h and "多选" in h and "判断" in h)):
            layout.type_i = i
            continue
        if layout.stem_i is None and "题干" in h:
            layout.stem_i = i
            continue
        opt_m = re.search(r"选项\s*([A-H])\b", h, re.IGNORECASE)
        if opt_m:
            letter = opt_m.group(1).upper()
            if letter not in layout.option_i:
                layout.option_i[letter] = i
            continue
        # 答案（A,B 英文逗号）等：以「答案」开头或含「答案」且非题干列
        if layout.answer_i is None and (
            hk.startswith("答案") or (re.search(r"答案", h) is not None and "题干" not in hk)
        ):
            layout.answer_i = i
            continue
        if layout.analysis_i is None and ("题目解析" in hk or ("解析" in hk and "答案" not in hk)):
            layout.analysis_i = i
            continue
        # 试题难度(简单，一般，困难) / 难易程度
        if layout.difficulty_i is None and (
            "试题难度" in hk
            or "难易程度" in hk
            or re.match(r"^难度", h)
            or ("难度" in hk and "知识点" not in hk and "试题难度" not in hk)
        ):
            layout.difficulty_i = i
            continue
        # 知识点ID(111,222 英文逗号)填写时 —— 与「知识点名称」区分
        if layout.kid_i is None and (
            "知识点ID" in hk or ("知识点" in hk and "ID" in hk and "名称" not in hk)
        ):
            layout.kid_i = i
            continue
        if layout.kname_i is None and ("知识点名称" in hk or ("知识点" in hk and "名称" in hk and "知识点ID" not in hk)):
            layout.kname_i = i
            continue
        # 标签(标签1,标签2 英文逗号)
        if layout.tag_i is None and hk.startswith("标签"):
            layout.tag_i = i
            continue
        # 指标(指标1,指标2 英文逗号)
        if layout.metric_i is None and hk.startswith("指标"):
            layout.metric_i = i
            continue
    if layout.tag_i is None:
        for i, raw in enumerate(columns):
            if _header_key(raw.strip()).startswith("标签"):
                layout.tag_i = i
                break
    return layout


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="解析Word试题并导出Excel")
    parser.add_argument("--input", required=True, help="Word文件路径（.docx）")
    parser.add_argument("--template", required=True, help="导题模板.xlsx路径")
    parser.add_argument("--output", required=True, help="输出Excel路径")
    parser.add_argument("--module", required=True, help="标签模块名称，如企业所得税")
    parser.add_argument(
        "--answer-separator",
        default="、",
        choices=[",", "，", "、", ""],
        help="答案多选分隔符，默认使用中文顿号",
    )
    return parser.parse_args()


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip())


def normalize_answer(answer_raw: str, separator: str = ",") -> str:
    normalized = answer_raw.strip()
    if normalized in {"正确", "错误"}:
        return normalized
    cleaned = re.sub(r"[^A-Za-z]", "", answer_raw)
    letters = [m.group(0).upper() for m in ANSWER_LETTER_RE.finditer(cleaned)]
    unique: List[str] = []
    for letter in letters:
        if letter not in unique:
            unique.append(letter)
    return separator.join(unique)


def answer_letter_count(answer_text: str) -> int:
    if answer_text in {"正确", "错误"}:
        return 1
    parts = [x for x in re.split(r"[,，、]", answer_text) if x]
    if parts:
        return len(parts)
    return len(re.findall(r"[A-H]", answer_text, re.IGNORECASE))


def normalize_difficulty(raw: str) -> str:
    text = re.sub(r"[，。；;、\s]+$", "", raw.strip())
    mapping = {
        "简单": "简单",
        "一般": "一般",
        "普通": "一般",
        "中等": "一般",
        "困难": "困难",
        "较难": "困难",
    }
    return mapping.get(text, "")


class Reader:
    def read_docx_paragraphs(self, docx_path: Path) -> List[str]:
        with ZipFile(docx_path) as docx:
            xml_content = docx.read("word/document.xml")
        root = ET.fromstring(xml_content)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        lines: List[str] = []
        for p in root.findall(".//w:p", ns):
            # itertext() 覆盖分段 w:t、修订等场景，比仅拼接 t.text 更稳
            line = "".join(p.itertext()).strip()
            if line:
                lines.append(line)
        return lines


class Parser:
    def parse_questions(self, paragraphs: List[str]) -> Tuple[List[Question], List[ParseError]]:
        questions: List[Question] = []
        errors: List[ParseError] = []
        current: Optional[Question] = None
        in_analysis = False

        for idx, line in enumerate(paragraphs):
            if ANSWER_KEY_SECTION_RE.search(line):
                break
            next_line = paragraphs[idx + 1] if idx + 1 < len(paragraphs) else ""
            q_match = QUESTION_START_RE.match(line)
            if q_match:
                if current:
                    questions.append(current)
                current = Question(number=int(q_match.group(1)))
                current.raw_lines.append(line)
                in_analysis = False
                stem = q_match.group(2).strip()
                if stem:
                    current.stem_lines.append(stem)
                continue
            if current is None:
                continue
            current.raw_lines.append(line)
            self._consume_line(current, line, in_analysis, errors)
            in_analysis = self._is_analysis_mode(in_analysis, line)
            if in_analysis and self._should_end_analysis(line, next_line):
                in_analysis = False
        if current:
            questions.append(current)
        return questions, errors

    def _should_end_analysis(self, current_line: str, next_line: str) -> bool:
        text = current_line.strip()
        if not text.endswith("。"):
            return False
        if not next_line:
            return True
        if ANSWER_KEY_SECTION_RE.search(next_line):
            return True
        if QUESTION_START_RE.match(next_line):
            return True
        if SECTION_BREAK_RE.match(next_line.strip()):
            return True
        return False

    def _consume_line(
        self,
        question: Question,
        line: str,
        in_analysis: bool,
        errors: List[ParseError],
    ) -> None:
        try:
            # 判断题特殊处理：无选项时直接出现答案，自动补全“正确/错误”选项
            if not question.answer_raw and not any(question.options.values()):
                if self._try_answer(question, line):
                    question.options["A"] = "正确"
                    question.options["B"] = "错误"
                    return
            if self._try_option(question, line, in_analysis):
                return
            if self._try_answer(question, line):
                return
            if self._try_analysis_start(question, line):
                return
            if in_analysis:
                question.analysis_lines.append(line)
            else:
                question.stem_lines.append(line)
        except Exception as exc:
            snippet = " | ".join(question.raw_lines[-3:])
            errors.append(ParseError(question.number, str(exc), snippet))
            print(f"[解析异常] 题号={question.number} 片段={snippet}")

    def _is_analysis_mode(self, current: bool, line: str) -> bool:
        if ANALYSIS_START_RE.match(line):
            return True
        return current

    def _try_option(self, question: Question, line: str, in_analysis: bool) -> bool:
        if in_analysis:
            return False
        inline_matches = list(INLINE_OPTION_RE.finditer(line))
        if len(inline_matches) >= 2:
            for match in inline_matches:
                key = match.group(1).upper()
                question.options[key] = normalize_text(match.group(2))
            return True
        match = OPTION_RE.match(line)
        if not match:
            return False
        key = match.group(1).upper()
        question.options[key] = normalize_text(match.group(2))
        return True

    def _try_answer(self, question: Question, line: str) -> bool:
        match = ANSWER_RE.match(line)
        if not match:
            return False
        content = normalize_text(match.group(1))
        head_match = ANSWER_HEAD_RE.match(content)
        if not head_match:
            question.answer_raw = content
            return True
        answer_part = normalize_text(head_match.group(1))
        question.answer_raw = answer_part
        remainder = normalize_text(content[head_match.end() :])
        if remainder:
            remainder = remainder.lstrip("：:，,;；。 ")
            if remainder:
                question.analysis_lines.append(remainder)
        return True

    def _try_analysis_start(self, question: Question, line: str) -> bool:
        match = ANALYSIS_START_RE.match(line)
        if not match:
            return False
        first = match.group(1).strip()
        if first:
            question.analysis_lines.append(first)
        return True


class Exporter:
    def __init__(self, answer_separator: str = "、") -> None:
        self.answer_separator = answer_separator

    def fill_row(self, layout: TemplateLayout, values: Dict[str, str]) -> List[str]:
        """按列下标生成一行，避免模板中存在空表头或重复列名时用 dict 键互相覆盖。"""
        n = len(layout.columns)
        row = [""] * n

        def put(idx: Optional[int], val: str) -> None:
            if idx is None or idx < 0 or idx >= n:
                return
            row[idx] = val

        put(layout.type_i, values.get("type", ""))
        put(layout.stem_i, values.get("stem", ""))
        put(layout.answer_i, values.get("answer", ""))
        put(layout.analysis_i, values.get("analysis", ""))
        put(layout.difficulty_i, values.get("difficulty", ""))
        put(layout.kid_i, values.get("kid", ""))
        put(layout.kname_i, values.get("kname", ""))
        put(layout.tag_i, values.get("tag", ""))
        put(layout.metric_i, values.get("metric", ""))
        for letter, idx in layout.option_i.items():
            put(idx, values.get(f"opt{letter}", ""))
        return row

    def empty_row(self, layout: TemplateLayout, module_name: str) -> List[str]:
        values: Dict[str, str] = {
            "type": "",
            "stem": "",
            "answer": "",
            "analysis": "",
            "difficulty": "",
            "kid": "",
            "kname": "",
            "tag": module_name,
            "metric": "",
        }
        for letter in OPTIONS:
            values[f"opt{letter}"] = ""
        return self.fill_row(layout, values)

    @staticmethod
    def row_as_dict(layout: TemplateLayout, row: List[str]) -> Dict[str, str]:
        """将按列下标的行转为列名→值（同名多列后者覆盖，仅用于测试/调试）。"""
        return {layout.columns[i]: row[i] if i < len(row) else "" for i in range(len(layout.columns))}

    def split_analysis_meta(self, analysis_text: str) -> Tuple[str, str, str]:
        difficulty = ""
        knowledge_name = ""
        # 先按「难度：/难易程度：」+ 标准档位提取，避免「困难知识点：」连写被整块吞掉
        diff_match = DIFFICULTY_LEVEL_RE.search(analysis_text)
        if diff_match:
            difficulty = normalize_difficulty(diff_match.group(1))
        k_match = KNOWLEDGE_RE.search(analysis_text)
        if k_match:
            knowledge_name = normalize_text(k_match.group(1))
        if not difficulty:
            # 仅在「知识点」之前的片段里兜底，减少误匹配正文里的「一般」等词
            head = analysis_text[: k_match.start()] if k_match else analysis_text
            for token in ["简单", "一般", "普通", "中等", "困难", "较难"]:
                if token in head:
                    difficulty = normalize_difficulty(token)
                    break
        cleaned = DIFFICULTY_LEVEL_RE.sub("", analysis_text)
        cleaned = KNOWLEDGE_RE.sub("", cleaned)
        return re.sub(r"\s+", " ", cleaned).strip(), difficulty, knowledge_name

    def detect_question_type(self, question: Question) -> str:
        answer = normalize_answer(question.answer_raw, separator=self.answer_separator)
        if answer in {"正确", "错误"}:
            return "3"
        filled_options = sum(1 for value in question.options.values() if value)
        if filled_options == 2 and question.options.get("A") == "正确" and question.options.get("B") == "错误":
            return "3"
        if answer_letter_count(answer) > 1:
            return "2"
        return "1"

    def split_by_type(self, questions: List[Question]) -> Dict[str, List[Question]]:
        grouped: Dict[str, List[Question]] = {"1": [], "2": [], "3": []}
        for question in questions:
            grouped[self.detect_question_type(question)].append(question)
        return grouped

    def build_row(self, question: Question, module_name: str, layout: TemplateLayout) -> List[str]:
        stem = normalize_text(" ".join(question.stem_lines))
        raw_analysis = normalize_text(" ".join(question.analysis_lines))
        analysis, difficulty, knowledge_name = self.split_analysis_meta(raw_analysis)
        answer = normalize_answer(question.answer_raw, separator=self.answer_separator)
        q_type = self.detect_question_type(question)
        option_values = {opt: question.options.get(opt, "") for opt in OPTIONS}
        if q_type == "3":
            if answer == "正确":
                answer = "A"
            elif answer == "错误":
                answer = "B"
            option_values = {opt: "" for opt in OPTIONS}

        values: Dict[str, str] = {
            "type": q_type,
            "stem": stem,
            "answer": answer,
            "analysis": analysis,
            "difficulty": difficulty,
            "kid": "",
            "kname": knowledge_name,
            "tag": module_name,
            "metric": "",
        }
        for letter in OPTIONS:
            values[f"opt{letter}"] = option_values.get(letter, "")
        return self.fill_row(layout, values)

    def build_rows_by_pipeline(
        self,
        questions: List[Question],
        module_name: str,
        layout: TemplateLayout,
    ) -> List[List[str]]:
        grouped = self.split_by_type(questions)
        ordered_questions = grouped["1"] + grouped["2"] + grouped["3"]
        return [self.build_row(q, module_name, layout) for q in ordered_questions]

    def export_excel(self, rows: List[List[str]], template_columns: List[str], output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.append(template_columns)
        n = len(template_columns)
        for row in rows:
            if len(row) < n:
                ws.append(row + [""] * (n - len(row)))
            elif len(row) > n:
                ws.append(row[:n])
            else:
                ws.append(row)
        wb.save(output_path)


def load_template_columns(template_path: Path) -> List[str]:
    wb = load_workbook(template_path, read_only=True)
    ws = wb.active
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    wb.close()
    if not first_row:
        raise ValueError("模板首行为空，无法识别列名")
    columns = [str(c) if c is not None else "" for c in first_row]
    # 1) 清理列名末尾空格；2) 删除模板末尾连续空列，避免出现 Unnamed: n
    columns = [c.rstrip() for c in columns]
    while columns and not columns[-1].strip():
        columns.pop()
    # 表头说明文案统一中文标点，避免出现 A,B 英文逗号提示
    columns = [c.replace("A,B", "A、B") for c in columns]
    return columns


class Validator:
    def _cell(self, row: List[str], layout: TemplateLayout, idx: Optional[int]) -> str:
        if idx is None or idx < 0:
            return ""
        if idx >= len(row):
            return ""
        return str(row[idx]).strip()

    def validate_all(
        self,
        rows: List[List[str]],
        questions: List[Question],
        module_name: str,
        layout: TemplateLayout,
    ) -> ValidationSummary:
        critical, warnings = 0, 0
        global_issues = self._check_missing_numbers(questions)
        critical += len(global_issues)
        for issue in global_issues:
            print(f"- {issue}")
        for question, row in zip(questions, rows):
            c_count, w_count = self._validate_question(question, row, module_name, layout)
            critical += c_count
            warnings += w_count
        total_issues = critical + warnings
        pass_rate = 1.0 if not rows else max(0.0, 1 - total_issues / max(len(rows), 1))
        return ValidationSummary(total=len(rows), critical=critical, warnings=warnings, pass_rate=pass_rate)

    def _check_missing_numbers(self, questions: List[Question]) -> List[str]:
        numbers = [q.number for q in questions]
        if not numbers:
            return []
        expected = set(range(min(numbers), max(numbers) + 1))
        missing = sorted(expected - set(numbers))
        if not missing:
            return []
        return [f"[全局] 题号不连续，缺失：{missing}"]

    def _validate_question(
        self,
        question: Question,
        row: List[str],
        module_name: str,
        layout: TemplateLayout,
    ) -> Tuple[int, int]:
        critical, warnings = 0, 0
        qid = question.number
        stem = self._cell(row, layout, layout.stem_i)
        if len(stem) < 10:
            warnings += 1
            print(f"- [题号{qid}] 题干过短（<10字符）")
        critical, warnings = self._check_options(row, qid, critical, warnings, layout)
        critical, warnings = self._check_answer(row, qid, critical, warnings, layout)
        analysis = self._cell(row, layout, layout.analysis_i)
        if len(analysis) < 20:
            warnings += 1
            print(f"- [题号{qid}] 解析缺失或过短（<20字符）")
        if not self._cell(row, layout, layout.difficulty_i):
            warnings += 1
            print(f"- [题号{qid}] 难度提取失败")
        tag_val = self._cell(row, layout, layout.tag_i)
        if not module_name.strip() or not tag_val:
            critical += 1
            print(f"- [题号{qid}] 标签未填写")
        return critical, warnings

    def _check_options(
        self,
        row: List[str],
        qid: int,
        critical: int,
        warnings: int,
        layout: TemplateLayout,
    ) -> Tuple[int, int]:
        q_type = self._cell(row, layout, layout.type_i)
        filled = 0
        for letter in OPTIONS:
            idx = layout.option_i.get(letter)
            if idx is not None and self._cell(row, layout, idx):
                filled += 1
        if q_type == "1" and filled < 2:
            warnings += 1
            print(f"- [题号{qid}] 单选题选项少于2个")
        return critical, warnings

    def _check_answer(
        self,
        row: List[str],
        qid: int,
        critical: int,
        warnings: int,
        layout: TemplateLayout,
    ) -> Tuple[int, int]:
        q_type = self._cell(row, layout, layout.type_i)
        answer = self._cell(row, layout, layout.answer_i)
        if q_type == "1" and answer_letter_count(answer) > 1:
            warnings += 1
            print(f"- [题号{qid}] 单选答案格式异常（出现多个字母）")
        return critical, warnings


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    template_path = Path(args.template)
    output_path = Path(args.output)
    module_name = args.module.strip()

    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")
    if input_path.suffix.lower() != ".docx":
        raise ValueError("当前版本仅支持 .docx 输入")
    if not template_path.exists():
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    reader = Reader()
    parser = Parser()
    exporter = Exporter(answer_separator=args.answer_separator)
    validator = Validator()

    print(f"开始解析Word: {input_path}")
    paragraphs = reader.read_docx_paragraphs(input_path)
    questions, parse_errors = parser.parse_questions(paragraphs)
    print(f"识别题目数量: {len(questions)}")
    for err in parse_errors:
        print(f"[解析异常] 题号={err.number} 片段={err.snippet}")

    template_columns = load_template_columns(template_path)
    template_layout = build_template_layout(template_columns)
    for msg in validate_template_layout(template_layout):
        print(msg)
    if template_layout.tag_i is None:
        print("警告: 未在模板首行识别到「标签」列，标签将无法写入，请检查表头文字。")
    typed_groups = exporter.split_by_type(questions)
    print(
        f"三阶段分流完成: 单选={len(typed_groups['1'])}, "
        f"多选={len(typed_groups['2'])}, 判断={len(typed_groups['3'])}"
    )
    ordered_questions = typed_groups["1"] + typed_groups["2"] + typed_groups["3"]

    rows: List[List[str]] = []
    for idx, question in enumerate(ordered_questions, start=1):
        try:
            rows.append(exporter.build_row(question, module_name, template_layout))
            if idx % 50 == 0 or idx == len(ordered_questions):
                print(f"进度: 已处理 {idx}/{len(ordered_questions)} 题")
        except Exception as exc:
            snippet = " | ".join(question.raw_lines[:3])
            print(f"[导出异常] 题号={question.number} 错误={exc} 片段={snippet}")
            rows.append(exporter.empty_row(template_layout, module_name))
    exporter.export_excel(rows, template_layout.columns, output_path)
    print(f"导出完成: {output_path}")

    print("\n=== 校验报告 ===")
    summary = validator.validate_all(rows, ordered_questions, module_name, template_layout)
    if parse_errors:
        summary.critical += len(parse_errors)
    print(summary.as_json())


if __name__ == "__main__":
    main()
